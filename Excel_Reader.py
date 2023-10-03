import openpyxl

def get_start_and_end_rows(sheet):
    total, start_row, end_row = [cell.row for cell in sheet['A'] if cell.value in ["Total", "REGIÓN"]]
    end_row =start_row+1
    #print("s row=",start_row,"e row=",end_row)
    return start_row, end_row

def get_start_and_end_columns(sheet):
    start_col, weekend_col_start = [cell.column for cell in sheet['11'] if cell.value in ["Lunes a Viernes", "Sábado-Domingo"]]
    end_col= sheet.max_column
    #print("s col =",start_col,"e col=", end_col,"w col=", weekend_col_start)
    return start_col, end_col, weekend_col_start

def get_time_stamp(sheet, start_row, start_col, end_col):
    time_stamp = []
    for row in range(start_row - 1, start_row):
        for col in range(start_col + 2, end_col + 1):
            value = sheet.cell(row, col).value
            if value not in ['Total', 'TOTAL']:
                time_stamp.append(value)
    #print(time_stamp)
    return time_stamp

def get_station_name(sheet, start_row, end_row, start_col):
    station_name = []
    for col in range(start_col - 1, start_col):
        for row in range(start_row, end_row):
            station_name.append(sheet.cell(row, col).value)
    return station_name

def creating_subsets(sheet, start_row, end_row, start_col, weekend_col_start, time_stamp, station_name):
    final_audience = []
    for val in [[start_col, weekend_col_start], [weekend_col_start, sheet.max_column + 1]]:
        if val[0] == start_col:
            flag = 0
        elif val[0] == weekend_col_start:
            flag = 1
        for row in range(start_row, end_row):
            audience = []
            dictionary = {}
            for col in range(val[0] + 2, val[1]):
                audience_value = sheet.cell(row, col).value
                if audience_value not in ['Total', 'TOTAL']:
                    audience.append(audience_value)
            for aud, time in zip(audience, time_stamp):
                new_time = f'{int(time.split(".")[0]):02}.00'
                if new_time in dictionary:
                    dictionary[new_time] += aud
                else:
                    dictionary[new_time] = aud
            final_audience.append({"station_name": station_name[row - start_row], "time": dictionary, "Flag": flag})
    return final_audience

def insert_data_into_database(final_audience):
    for i in final_audience:
        for range_data in range(0, 5) if i["Flag"] == 0 else range(5, 7):
            op = f'Insert into radio_audience(station_name, timestamp, audience, flag) values ' \
                 f'({i["station_name"]}, {i["time"]}, {i["Flag"]});'
            return(op)

if __name__ == "__main__":
    wb = openpyxl.load_workbook('spain_data.xlsx')
    sheet = wb['Hoja2']
    
    start_row, end_row = get_start_and_end_rows(sheet)
    start_col, end_col, weekend_col_start = get_start_and_end_columns(sheet)
    time_stamp = get_time_stamp(sheet, start_row, start_col, end_col)
    station_name = get_station_name(sheet, start_row, end_row, start_col)
    final_audience =creating_subsets(sheet, start_row, end_row, start_col, weekend_col_start, time_stamp, station_name)
    
    insert_data_into_database(final_audience)
